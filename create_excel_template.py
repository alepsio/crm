import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# Crea un nuovo workbook usando pandas per assicurarsi che il formato sia corretto
columns = [
    "Nome", "Cognome", "Codice Fiscale", "Email", "Cellulare", 
    "Partita IVA", "PEC", "Codice Univoco", "Residenza", 
    "Domicilio Fiscale", "Attività", "Data Nascita", "Luogo Nascita"
]

# Crea un DataFrame vuoto con le colonne corrette
df = pd.DataFrame(columns=columns)

# Aggiungi dati di esempio
df.loc[0] = [
    "Mario", "Rossi", "RSSMRA80A01H501U", "mario.rossi@example.com", "3331234567", 
    "12345678901", "mario.rossi@pec.it", "ABC123", "Via Roma 1, Roma", 
    "Via Roma 1, Roma", "Commercialista", "01/01/1980", "Roma"
]
df.loc[1] = [
    "Anna", "Bianchi", "BNCNNA85B02H501V", "anna.bianchi@example.com", "3339876543", 
    "", "anna.bianchi@pec.it", "", "Via Milano 2, Milano", 
    "Via Milano 2, Milano", "Avvocato", "02/02/1985", "Milano"
]

# Salva il DataFrame in Excel
file_path = "app/static/templates/template_importazione_clienti.xlsx"
os.makedirs(os.path.dirname(file_path), exist_ok=True)
df.to_excel(file_path, index=False)

# Ora apri il file con openpyxl per formattarlo
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Stili
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# Formatta le intestazioni
for col_idx, column in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border
    
    # Imposta la larghezza della colonna
    ws.column_dimensions[get_column_letter(col_idx)].width = 20

# Formatta i dati di esempio
for row_idx in range(2, 4):  # 2 righe di dati
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(vertical="center")

# Aggiungi note informative
ws.cell(row=5, column=1, value="Note:").font = Font(bold=True)
ws.cell(row=6, column=1, value="- I campi Nome, Cognome, Codice Fiscale ed Email sono obbligatori")
ws.cell(row=7, column=1, value="- Per le date utilizzare il formato GG/MM/AAAA")
ws.cell(row=8, column=1, value="- Rimuovere le righe di esempio prima dell'importazione")
ws.cell(row=9, column=1, value="- Limiti di lunghezza: Partita IVA (11), Codice Fiscale (16), Codice Univoco (7)")

# Crea la directory se non esiste
os.makedirs("app/static/templates", exist_ok=True)

# Salva il file
file_path = "app/static/templates/template_importazione_clienti.xlsx"
wb.save(file_path)

print(f"Template Excel creato con successo: {file_path}")